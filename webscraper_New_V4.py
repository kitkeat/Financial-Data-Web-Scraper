####################################################################################
######## Web Scrapper ##############################################################
####################################################################################
# V3 - Display Data on Python in table format via pandas
# V4 - Scrap Financial statement

#--------------------------- Load Excel Workbook ----------------------------------#
# https://realpython.com/openpyxl-excel-spreadsheets-python/#getting-started-with-openpyxl
from openpyxl import load_workbook
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series)
import os
import time
import shutil
import sys
import pandas as pd
from bs4 import BeautifulSoup as soup
import requests
import bs4

from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException

current_dir = os.path.dirname(os.path.abspath(__file__))
os.chdir(current_dir)
os.getcwd()
print(os.getcwd())

# Request user to enter ticker
ticker = input('Please input the ticker: ')
ticker = ticker.upper()

# Create copy of excel sheet #
curr_month = str(time.localtime().tm_mon)
curr_day = str(time.localtime().tm_mday)
curr_hour = str(time.localtime().tm_hour)
curr_min = str(time.localtime().tm_min)
curr_sec = str(time.localtime().tm_sec)
print("Current month: " + str(curr_month))
print("Current day: " + str(curr_day))
print("Current hour: " + str(curr_hour))
print("Current min: " + str(curr_min))
source = "Valuation.xlsx"
destination = "Valuation_" + curr_month + "_" + curr_day + "_" + curr_hour + "_" + curr_min + "_" + curr_sec + ".xlsx"
shutil.copyfile(source, destination)

# Load workbook without formula, will not be saved. Just for reading data
workbook = load_workbook(filename="Valuation.xlsx",data_only=True)

# Get all worksheet names and populate inside the array
# Check the existence of requested ticker within excel sheet
index = 0
ArrSize=1000
arr_worksheet = [None]*ArrSize
get_sheet = workbook.sheetnames
print('Worksheet names:')
for i in get_sheet:
    arr_worksheet[index] = i
    index = index + 1
    print(str(index)+ ': ' + i)
    if ticker == i:
        print("Ticker (" + str(i) + ") already exist in excel sheet")
        input("Press any key to exit")
        sys.exit()

# Toggle between workbook as workaround for bug
workbook.active = workbook["Summary"]
active_sheet = workbook.active
time.sleep(1)
workbook.active = workbook["Template"]
active_sheet = workbook.active
time.sleep(1)
workbook.active = workbook["Summary"]
active_sheet = workbook.active
time.sleep(1)

# Create a list with initial capacity
Size = 10
x_axis = [None]*Size #year
y_axis_debt_equity_ratio = [None]*Size 
y_axis_net_income = [None]*Size 
y_axis_bookvalue_pershare = [None]*Size 
y_axis_debt_equity_ratio = [None]*Size

w, h = 18, 10
keyratios_data_2D_arr = [[0 for x in range(w)] for y in range(h)] 
data_edited_2D_arr = [[0 for x in range(w)] for y in range(h)]

#-----------------------------------------------#
#---------------- Summary Sheet ----------------#
# Set Summary as active sheet
cell_bookmark = 1
cell_no = 2
cell_ticker = 3
cell_marketprice = 4
cell_intrinsicvalue = 5
cell_PE_Ratio = 6
cell_Beta = 7
cell_comment = 8
cell_totalnoofstock = 9

# Check total no of stock in the summary page
if active_sheet.cell(row=1,column=cell_totalnoofstock).value == 'Total No of Stocks':
    summary_index = active_sheet.cell(row=2,column=cell_totalnoofstock).value
    print('Total No of Stocks: ' + str(summary_index))
else:
    print('Error: Reading wrong cell to get "Total No of Stocks"')    
summary_index_int = int(summary_index) + 2
active_sheet.cell(row=summary_index_int,column=cell_ticker).value = str(ticker)
workbook.close()

# Load workbook second time to include formula
workbook = load_workbook(filename="Valuation.xlsx")
workbook.active = workbook["Summary"]
active_sheet = workbook.active

#https://www.promptcloud.com/blog/how-to-scrape-yahoo-finance-data-using-python/


""" my_url = requests.get('https://finance.yahoo.com/quote/' + ticker + '/')
page_soup = soup(my_url.content.decode('utf-8'), "html.parser")
 """
url = 'https://finance.yahoo.com/quote/' + ticker + '/'
browser = webdriver.Chrome(current_dir + '\webdriver\chromedriver')
browser.get(url)
delay = 10
try:
        myElem = WebDriverWait(browser, delay).until(EC.presence_of_element_located((By.XPATH, "//div[contains(text(),'Revenue')]")))
        print( "Page is ready!")
except TimeoutException:
        print ("Loading took too much time!")
html = browser.page_source
soup = bs4.BeautifulSoup(html, 'html.parser')
#-------------------------------------------------------------------------------------------#
#------------------------- Load Yahoo Finance page to grab current data --------------------#
#-------------------------------------------------------------------------------------------#

# Get Market Price
#market_price_data = page_soup.find('span',attrs={'class':'Trsdu(0.3s) Fw(b) Fz(36px) Mb(-4px) D(ib)'})   #get market price from Yahoo!
market_price_data = soup.find('div', attrs={'class':'D(ib) Mend(20px)'})
print(market_price_data)
for child in market_price_data.contents:
    #print(child.text)
    data_cleaned = child.text.replace(',', '')
    market_price = float(data_cleaned)
    print(ticker + ' market price: $' + data_cleaned)
    break #read the first data and then end the loop
active_sheet.cell(row=summary_index_int,column=cell_marketprice).value = market_price

# Get beta value
beta_data = soup.find('td',{'data-test': 'BETA_5Y-value'})
beta_text = beta_data.text
if beta_text == 'N/A':
    beta = str(beta_text)
else:
    beta = float(beta_text)
active_sheet.cell(row=summary_index_int,column=cell_Beta).value = beta
print(beta)

# Get PE Ratio
pe_ratio_data = soup.find('td',{'data-test': 'PE_RATIO-value'})
pe_ratio_text = pe_ratio_data.text
if pe_ratio_text == 'N/A':
    pe_ratio = str(pe_ratio_text)
else:
    pe_ratio = float(pe_ratio_text)
active_sheet.cell(row=summary_index_int,column=cell_PE_Ratio).value = pe_ratio
print(pe_ratio)

#-------------------------------------------------------------------------------------------#
#------------------------- Load MorningStar page to grab past data -------------------------#
#-------------------------------------------------------------------------------------------#

workbook.active = workbook["Template"]
sheet = workbook["Template"]
workbook.copy_worksheet(sheet)
new_sheet = workbook['Template Copy']
new_sheet.title = ticker
workbook.active = workbook[ticker]
active_sheet = workbook.active

url = 'http://financials.morningstar.com/ratios/r.html?t=' + ticker
browser = webdriver.Chrome(current_dir + '\webdriver\chromedriver')
browser.get(url)
delay = 10
try:
        myElem = WebDriverWait(browser, delay).until(EC.presence_of_element_located((By.XPATH, "//div[contains(text(),'Revenue')]")))
        print( "Page is ready!")
except TimeoutException:
        print ("Loading took too much time!")
html = browser.page_source
soup = bs4.BeautifulSoup(html, 'html.parser')

# Title
title = soup.find('div', {'class':'r_title'})
print(title.h1.text)
active_sheet.cell(row=2,column=1).value = title.h1.text
active_sheet.cell(row=2,column=2).value = ticker

# Get Data Year
for x in range(10):
    concat1 = 'Y' + str(x)
    year = soup.find('th', {'id':concat1})
    idx = 0
    while year == None and idx < 5:
        year = soup.find('th', {'id':concat1})
        idx = idx + 1  
        print("Repeat")         
    print(year.text)
    year_str = year.text
    year_str = year_str[:4]
    year_float = float(year_str)
    keyratios_data_2D_arr[x][0] = year_float
    ##active_sheet.cell(row=4,column=x+2).value = year_float

# Get Data
for y in range(12):
    id = 'i' + str(y)
    revenue = soup.find('th', {'id':id})   
    print(revenue.text)
    for x in range(10):
        concat1 = 'Y' + str(x) + ' ' + id
        values = soup.find('td', {'headers':concat1})
        values1 = values.text
        if values1 == "—":
            values1 = "0"
        values1 = values1.replace(',', '')
        values_float = float(values1)
        if y == 4: #net income
            y_axis_net_income[x] = values_float
            
        elif y == 8: #book value per share
            y_axis_bookvalue_pershare[x] = values_float
            
            
        if y >= 7:
            ##active_sheet.cell(row=y+6,column=x+2).value = values_float
            if y == 7:
                keyratios_data_2D_arr[x][9] = values_float
            elif y == 8:
                keyratios_data_2D_arr[x][10] = values_float
            elif y == 9:
                keyratios_data_2D_arr[x][11] = values_float
            elif y == 10:
                keyratios_data_2D_arr[x][12] = values_float
            elif y == 11:
                keyratios_data_2D_arr[x][13] = values_float
        else:
            ##active_sheet.cell(row=y+5,column=x+2).value = values_float
            if y == 0:
                keyratios_data_2D_arr[x][1] = values_float
            elif y == 1:
                keyratios_data_2D_arr[x][2] = values_float
            elif y == 2:
                keyratios_data_2D_arr[x][3] = values_float
            elif y == 3:
                keyratios_data_2D_arr[x][4] = values_float
            elif y == 4:
                keyratios_data_2D_arr[x][5] = values_float
            elif y == 5:
                keyratios_data_2D_arr[x][6] = values_float
            elif y == 6:
                keyratios_data_2D_arr[x][7] = values_float
        print(values.text)

# Free cashflow Per Share
id = 'i' + '90'
revenue = soup.find('th', {'id':id})   
print(revenue.text)
for x in range(10):
    concat1 = 'Y' + str(x) + ' ' + id
    values = soup.find('td', {'headers':concat1}) 
    values1 = values.text
    if values1 == "—":
        values1 = "0"
    values1 = values1.replace(',', '')
    values_float = float(values1)
    ##active_sheet.cell(row=18,column=x+2).value = values_float
    keyratios_data_2D_arr[x][14] = values_float
    print(values.text)

# Payout ratio
id = 'i' + '91'
revenue = soup.find('th', {'id':id})   
print(revenue.text)
for x in range(10):
    concat1 = 'Y' + str(x) + ' ' + id
    values = soup.find('td', {'headers':concat1}) 
    values1 = values.text
    if values1 == "—":
        values1 = "0"
    values1 = values1.replace(',', '')
    values_float = float(values1)
    ##active_sheet.cell(row=12,column=x+2).value = values_float
    keyratios_data_2D_arr[x][8] = values_float
    print(values.text)

# Working Capital
id = 'i' + '80'
revenue = soup.find('th', {'id':id})   
print(revenue.text)
for x in range(10):
    concat1 = 'Y' + str(x) + ' ' + id
    values = soup.find('td', {'headers':concat1})
    values1 = values.text
    if values1 == "—":
        values1 = "0"
    values1 = values1.replace(',', '')
    values_float = float(values1)
    ##active_sheet.cell(row=19,column=x+2).value = values_float
    keyratios_data_2D_arr[x][15] = values_float
    print(values.text)

# Current Ratio
id = 'i65'
idx = 0
curr_ratio = soup.find('th', {'id':id}) 
while curr_ratio == None and idx < 40:
    curr_ratio = soup.find('th', {'id':id}) 
    idx = idx + 1  
    print("Repeat")         
print(curr_ratio.text)

for x in range(10):
    concat1 = 'lfh-Y' + str(x) + ' lfh-liquidity ' + id
    values = soup.find('td', {'headers':concat1})
    values1 = values.text
    if values1 == "—":
        values1 = "0"
    values1 = values1.replace(',', '')
    values_float = float(values1)
    ##active_sheet.cell(row=20,column=x+2).value = values_float
    keyratios_data_2D_arr[x][16] = values_float
    print(values.text)     

# Debt/Equity Ratio
id = 'i68'
revenue = soup.find('th', {'id':id})   
print(revenue.text)
for x in range(10):
    concat1 = 'lfh-Y' + str(x) + ' lfh-liquidity ' + id
    values = soup.find('td', {'headers':concat1})
    values1 = values.text
    if values1 == "—":
        values1 = "0"
    values1 = values1.replace(',', '')
    values_float = float(values1)
    y_axis_debt_equity_ratio[x] = values_float
    ##active_sheet.cell(row=21,column=x+2).value = values_float
    keyratios_data_2D_arr[x][17] = values_float
    print(values.text)  


#Construct table in console
pd.set_option('display.max_columns', None)
pd.set_option('display.width', None)
pd.set_option('display.max_colwidth', -1)
df = pd.DataFrame(keyratios_data_2D_arr, columns= ['Year','Revenue','Gross Margin','Operating Income','Operating Margin','Net Income','Earnings Per Share','Dividends Per Share','Payout Ratio','No. of Shares','Book Value Per Share','Operating Cash Flow','Capital Spending','Free Cash Flow','Free Cash Flow Per Share','Working Capital','Current Ratio','Debt/Equity'])
df_transposed = df.T
print(df_transposed)

################################################################################################################################################
############### Change change to Income Statement page #########################################################################################
################################################################################################################################################

browser.find_element_by_xpath("//a[normalize-space()='Financials']").click()
try:
        myElem = WebDriverWait(browser, delay).until(EC.presence_of_element_located((By.XPATH, "//div[contains(text(),'Revenue')]")))
        print( "Page is ready!")
except TimeoutException:
        print ("Loading took too much time!")
html = browser.page_source
soup = bs4.BeautifulSoup(html, 'html.parser')

w, h = 22, 5
incomestatement_data_2D_arr = [[0 for x in range(w)] for y in range(h)] 
incomestatement_data_edited_2D_arr = [[0 for x in range(w)] for y in range(h)]

#Year
print('Year')
for h in range(1,6):
        id = 'Y_' + str(h)
        data = soup.find('div', {'id':id})
        try:
                print(data.text)
                data_cleaned = data.text[:4]
                incomestatement_data_2D_arr[0][h-1] = data_cleaned
        except:
                ('Error encountered')

for index in range(1,21):
        if index == 1:
                title = 'Revenue'
                y = 'i1'
        elif index == 2:
                title = 'Cost of Revenue'
                y = 'i6'
        elif index == 3:
                title = 'Gross Profit'
                y = 'i10'
        elif index == 4:
                title = 'Research and development'
                y = 'i11'
        elif index == 5:
                title = 'Sales, General and Administrative'
                y = 'i12'
        elif index == 6:
                title = 'Restructuring, Merger and Acquisition'
                y = 'i25'
        elif index == 7:
                title = 'Other operating expense'
                y = 'i29'
        elif index == 8:
                title = 'Total Operating Expense'
                y = 'ttg3'
        elif index == 9:
                title = 'Operating Income'
                y = 'i30'
        elif index == 10:
                title = 'Interest Expense'
                y = 'i51'
        elif index == 11:
                title = 'Other Income (Expense)'
                y = 'i52'
        elif index == 12:
                title = 'Income before taxes'
                y = 'i60'
        elif index == 13:
                title = 'Provision for income taxes'
                y = 'i61'
        elif index == 14:
                title = 'Net income from continuing operations'
                y = 'i70'
        elif index == 15:
                title = 'Net Income'
                y = 'i80'
        elif index == 16:
                title = 'Net income available to common stockholders '
                y = 'i82'
        elif index == 17:
                title = 'Earnings per share - Basic'
                y = 'i83'
        elif index == 18:
                title = 'Earnings per share - Diluted'
                y = 'i84'
        elif index == 19:
                title = 'Weighted Average Shareholder Outstanding - Basic'
                y = 'i85'
        elif index == 20:
                title = 'Weighted Average Shareholder Outstanding - Diluted'
                y = 'i86'
        elif index == 21:
                title = 'EBITDA'
                y = 'i90'

        print(title)
        id = 'data_' + y
        data = soup.find('div', {'id':id})
        data = data.div
        for a in range(1,6):
                print(data.text)
                data_cleaned = data.text.replace(',', '')
                data_cleaned = data_cleaned.replace('—', '0')
                data_cleaned = data_cleaned.replace('(', '-')
                data_cleaned = data_cleaned.replace(')', '')
                incomestatement_data_2D_arr[a-1][index] = data_cleaned
                data = data.next_sibling

################################################################################################################################################
############### Change change to Cash Flow page #########################################################################################
################################################################################################################################################

browser.find_element_by_xpath("//h2[normalize-space()='Cash Flow']").click()
try:
        myElem = WebDriverWait(browser, delay).until(EC.presence_of_element_located((By.XPATH, "//div[@id='data_i1']//div[@id='Y_1']")))
        print( "Page is ready!")
except TimeoutException:
        print ("Loading took too much time!")
html = browser.page_source
soup = bs4.BeautifulSoup(html, 'html.parser')

w, h = 51, 5
cashflow_data_2D_arr = [[0 for x in range(w)] for y in range(h)] 
cashflow_data_edited_2D_arr = [[0 for x in range(w)] for y in range(h)]

#Year
print('Year')
for h in range(1,6):
        id = 'Y_' + str(h)
        data = soup.find('div', {'id':id})
        try:
                print(data.text)
                data_cleaned = data.text[:4]
                cashflow_data_2D_arr[0][h-1] = data_cleaned
        except:
                ('Error encountered')

for index in range(1,51):
        if index == 1:
                title = 'Net Income'
                y = 'i1'
        elif index == 2:
                title = 'Depreciation and Amortization'
                y = 'i2'
        elif index == 3:
                title = 'Amortization of debt discount/premium and issuance costs'
                y = 'i3'
        elif index == 4:
                title = 'Investment/asset impairment charges'
                y = 'i4'
        elif index == 5:
                title = 'Investments losses (gains)'
                y = 'i5'
        elif index == 6:
                title = 'Deferred income taxes'
                y = 'i6'
        elif index == 7:
                title = '(Gain) Loss from discontinued operations'
                y = 'i7'
        elif index == 8:
                title = 'Extraordinary items'
                y = 'i8'
        elif index == 9:
                title = 'Cumulative effect of accounting change'
                y = 'i9'
        elif index == 10:
                title = 'Stock based compensation'
                y = 'i10'
        elif index == 11:
                title = 'Change in working capital'
                y = 'i15'
        elif index == 12:
                title = 'Accounts receivable'
                y = 'i16'
        elif index == 13:
                title = 'Inventory'
                y = 'i17'
        elif index == 14:
                title = 'Prepaid expenses'
                y = 'i18'
        elif index == 15:
                title = 'Accounts payable'
                y = 'i19'
        elif index == 16:
                title = 'Accrued liabilities'
                y = 'i20'
        elif index == 17:
                title = 'Interest payable'
                y = 'i21'
        elif index == 18:
                title = 'Income taxes payable'
                y = 'i22'
        elif index == 19:
                title = 'Other working capital'
                y = 'i23'
        elif index == 20:
                title = 'Other non-cash items'
                y = 'i30'
        elif index == 21:
                title = 'Net cash provided by operating activities'
                y = 'tts1'
        elif index == 22:
                title = 'Investments in property, plant, and equipment'
                y = 'i31'
        elif index == 23:
                title = 'Property, plant, and equipment reductions'
                y = 'i32'
        elif index == 24:
                title = 'Acquisitions, net'
                y = 'i33'
        elif index == 25:
                title = 'Purchases of investments'
                y = 'i34'
        elif index == 26:
                title = 'Sales/Maturities of investments'
                y = 'i35'
        elif index == 27:
                title = 'Investments in technologies'
                y = 'i36'
        elif index == 28:
                title = 'Sales of technologies'
                y = 'i37'
        elif index == 29:
                title = 'Purchases of intangibles'
                y = 'i38'
        elif index == 30:
                title = 'Sales of intangibles'
                y = 'i39'
        elif index == 31:
                title = 'Other investing activities'
                y = 'i60'
        elif index == 32:
                title = 'Net cash used for investing activities'
                y = 'tts2'
        elif index == 33:
                title = 'Debt issued'
                y = 'i61'
        elif index == 34:
                title = 'Debt repayment'
                y = 'i62'
        elif index == 35:
                title = 'Preferred stock issued'
                y = 'i63'
        elif index == 36:
                title = 'Preferred stock repaid'
                y = 'i64'
        elif index == 37:
                title = 'Warrant issued'
                y = 'i65'
        elif index == 38:
                title = 'Common stock issued'
                y = 'i66'
        elif index == 39:
                title = 'Common stock repurchased'
                y = 'i67'
        elif index == 40:
                title = 'Excess tax benefit from stock based compensation'
                y = 'i68'
        elif index == 41:
                title = 'Dividend paid'
                y = 'i69'
        elif index == 42:
                title = 'Dividend payable'
                y = 'i70'
        elif index == 43:
                title = 'Other financing activities'
                y = 'i90'
        elif index == 44:
                title = 'Net cash provided by (used for) financing activities'
                y = 'tts3'
        elif index == 45:
                title = 'Effect of exchange rate changes'
                y = 'i91'
        elif index == 46:
                title = 'Net change in cash'
                y = 'i93'
        elif index == 47:
                title = 'Cash at beginning of period'
                y = 'i94'
        elif index == 48:
                title = 'Cash at end of period'
                y = 'i95'
        elif index == 49:
                title = 'Operating cash flow'
                y = 'i100'
        elif index == 50:
                title = 'Capital expenditure'
                y = 'i96'
        elif index == 51:
                title = 'Free cash flow'
                y = 'i97'

        print(title)
        id = 'data_' + y
        data = soup.find('div', {'id':id})
        data = data.div
        for a in range(1,6):
                print(data.text)
                data_cleaned = data.text.replace(',', '')
                data_cleaned = data_cleaned.replace('—', '0')
                data_cleaned = data_cleaned.replace('(', '-')
                data_cleaned = data_cleaned.replace(')', '')
                cashflow_data_2D_arr[a-1][index] = data_cleaned
                data = data.next_sibling

browser.close()

######################################################################################################################################
############################### Excel Data Input #####################################################################################
######################################################################################################################################

""" #----------------------------- Add Charts -------------------------------------------------------------------------------------------#
# Net Income Chart
chart1 = ScatterChart()
chart1.title = "Net Income"
chart1.style = 13
chart1.legend = None
chart1.width = 10
xvalues = Reference(active_sheet, min_col=2, min_row=4, max_col=11)
values = Reference(active_sheet, min_col=2, min_row=9, max_col=11)
series = Series(values, xvalues, title_from_data=False)
chart1.series.append(series)
active_sheet.add_chart(chart1, "A35") # Location of the chart

# Debt/Equity Ratio Chart
chart2 = ScatterChart()
chart2.title = "Debt/Equity Ratio"
chart2.style = 13
chart2.legend = None
chart2.width = 10
xvalues = Reference(active_sheet, min_col=2, min_row=4, max_col=11)
values = Reference(active_sheet, min_col=2, min_row=21, max_col=11)
series = Series(values, xvalues, title_from_data=False)
chart2.series.append(series)
active_sheet.add_chart(chart2, "D35") # Location of the chart

# Book Value Per Share Chart
chart3 = ScatterChart()
chart3.title = "Book Value Per Share"
chart3.style = 13
chart3.legend = None
chart3.width = 10
xvalues = Reference(active_sheet, min_col=2, min_row=4, max_col=11)
values = Reference(active_sheet, min_col=2, min_row=14, max_col=11)
series = Series(values, xvalues, title_from_data=False)
chart3.series.append(series)
active_sheet.add_chart(chart3, "H35") # Location of the chart """


#-------------------------------- Append Data in specific sheets ------------------------------------------------------------------#
# Add link data from summary page to specific sheet
ticker_celllocation_summary = str(summary_index_int)
active_sheet.cell(row=94,column=2).value = '=Summary!G'+ticker_celllocation_summary # Beta
active_sheet.cell(row=95,column=2).value = '=Summary!F'+ticker_celllocation_summary # PE Ratio
active_sheet.cell(row=96,column=2).value = '=Summary!D'+ticker_celllocation_summary # Market Price

# Add key ratio data into specific sheet
for y in range(18):
    for x in range(0,10):
        active_sheet.cell(row=y+4,column=x+2).value = keyratios_data_2D_arr[x][y]

# Check data starting year
for x in range (0,10):
    if keyratios_data_2D_arr[x][0] == float(incomestatement_data_2D_arr[0][0]):
        incomeStatement_start_pos = x
        break
for x in range (0,10):
    if keyratios_data_2D_arr[x][0] == float(cashflow_data_2D_arr[0][0]):
        cashflow_start_pos = x
        break

# Add income statement data into specific sheet
for y in range(1,21):
    for x in range(0,5):
        active_sheet.cell(row=y+21,column=x+2+incomeStatement_start_pos).value = float(incomestatement_data_2D_arr[x][y])

# Add income statement data into specific sheet
for y in range(1,51):
    for x in range(0,5):
        active_sheet.cell(row=y+41,column=x+2+cashflow_start_pos).value = float(cashflow_data_2D_arr[x][y])

# Open summary sheet
workbook.active = workbook["Summary"]
active_sheet = workbook.active

# Updata summary page 
active_sheet.cell(row=int(ticker_celllocation_summary),column=cell_intrinsicvalue).value = '='+ ticker +'!B122' # Intrinsic value
active_sheet.cell(row=int(ticker_celllocation_summary),column=cell_ticker).value = ticker # Ticker
active_sheet.cell(row=2,column=cell_totalnoofstock).value = summary_index + 1 # 

# Save and close excel file
workbook.save("Valuation.xlsx")
workbook.close()

#-------------------------- The End -----------------------------------------------------------------------------------------------#
sys.exit()
